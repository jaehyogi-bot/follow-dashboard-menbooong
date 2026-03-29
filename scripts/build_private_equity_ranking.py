from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADERS = {"User-Agent": "Mozilla/5.0"}

INVESTOR_CODES = {
    "private": ("3100", "private_net_buy_bil_krw"),
    "trust": ("3000", "trust_net_buy_bil_krw"),
    "institution": ("7050", "institution_net_buy_bil_krw"),
    "foreigner": ("9000", "foreigner_net_buy_bil_krw"),
}

MARKET_TO_KRX = {
    "ALL": "ALL",
    "KOSPI": "STK",
    "KOSDAQ": "KSQ",
    "KONEX": "KNX",
}

NAVER_MARKET_MAP = {
    "KOSPI": "0",
    "KOSDAQ": "1",
}

MIN_MARKET_CAP_BIL = 3000.0
MIN_TURNOVER_BIL = 30.0
MAX_ABS_CHANGE_PCT = 25.0

MAIN_WEIGHT = 0.85
INST_WEIGHT = 0.10
FOREIGN_WEIGHT = 0.05


def parse_int(value: str) -> int:
    cleaned = (value or "").replace(",", "").replace("%", "").replace("-", "0").strip()
    return int(cleaned or "0")


def parse_float(value: str) -> float:
    cleaned = (value or "").replace(",", "").replace("%", "").strip()
    if cleaned in {"", "-"}:
        return 0.0
    return float(cleaned)


def to_billion_krw(series: pd.Series) -> pd.Series:
    return (series.astype("float64") / 100_000_000).round(4)


def get_recent_trading_window(session: requests.Session, anchor_date: str | None = None) -> tuple[str, str]:
    unique_dates: list[str] = []

    for page in range(1, 201):
        response = session.get(
            f"https://finance.naver.com/item/sise_day.naver?code=005930&page={page}",
            headers={**HEADERS, "Referer": "https://finance.naver.com/"},
            timeout=30,
        )
        response.raise_for_status()
        page_dates = re.findall(r"<span class=\"tah p10 gray03\">(\d{4}\.\d{2}\.\d{2})</span>", response.text)
        for value in page_dates:
            compact = value.replace(".", "")
            if anchor_date and compact > anchor_date:
                continue
            if compact not in unique_dates:
                unique_dates.append(compact)

        if len(unique_dates) >= 5:
            break

    if len(unique_dates) < 5:
        if anchor_date:
            raise RuntimeError(f"Failed to detect 5 trading days up to anchor date {anchor_date}.")
        raise RuntimeError("Failed to detect the recent 5 trading days.")

    return unique_dates[4], unique_dates[0]


def fetch_krx_investor_frame(
    session: requests.Session,
    start: str,
    end: str,
    market: str,
    investor_code: str,
    value_column: str,
) -> pd.DataFrame:
    response = session.post(
        "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd",
        headers={
            **HEADERS,
            "Referer": "https://data.krx.co.kr/contents/MDC/MDI/outerLoader/index.cmd?screenId=MDCSTAT024",
        },
        data={
            "bld": "dbms/MDC_OUT/STAT/standard/MDCSTAT02401_OUT",
            "locale": "ko_KR",
            "strtDd": start,
            "endDd": end,
            "mktId": MARKET_TO_KRX[market],
            "invstTpCd": investor_code,
            "share": "1",
            "money": "1",
            "csvxls_isNo": "false",
        },
        timeout=30,
    )
    response.raise_for_status()
    rows = response.json().get("output", [])
    if not rows:
        raise RuntimeError(
            f"No KRX investor data returned for investor={investor_code}, market={market}, period={start}-{end}."
        )

    frame = pd.DataFrame(rows).rename(
        columns={"ISU_SRT_CD": "code", "ISU_NM": "name", "NETBID_TRDVAL": value_column}
    )
    frame["code"] = frame["code"].astype(str).str.zfill(6)
    frame[value_column] = frame[value_column].map(parse_int)
    return frame[["code", "name", value_column]]


def fetch_krx_snapshot(
    session: requests.Session,
    start: str,
    end: str,
    market: str,
    investor_code: str,
) -> dict[str, int]:
    response = session.post(
        "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd",
        headers={
            **HEADERS,
            "Referer": "https://data.krx.co.kr/contents/MDC/MDI/outerLoader/index.cmd?screenId=MDCSTAT024",
        },
        data={
            "bld": "dbms/MDC_OUT/STAT/standard/MDCSTAT02401_OUT",
            "locale": "ko_KR",
            "strtDd": start,
            "endDd": end,
            "mktId": MARKET_TO_KRX[market],
            "invstTpCd": investor_code,
            "share": "1",
            "money": "1",
            "csvxls_isNo": "false",
        },
        timeout=30,
    )
    response.raise_for_status()
    return {
        str(row["ISU_SRT_CD"]).zfill(6): parse_int(row.get("NETBID_TRDVAL", "0"))
        for row in response.json().get("output", [])
    }


def get_naver_last_page(session: requests.Session, sosok: str) -> int:
    response = session.get(
        f"https://finance.naver.com/sise/sise_market_sum.naver?sosok={sosok}&page=1",
        headers={**HEADERS, "Referer": "https://finance.naver.com/"},
        timeout=30,
    )
    response.raise_for_status()
    html = response.text
    match = re.search(r'page=(\d+)".*?\ub9e8\ub4a4', html, re.DOTALL)
    if match:
        return int(match.group(1))
    pages = [int(page) for page in re.findall(r"page=(\d+)", html)]
    return max(pages) if pages else 1


def fetch_naver_market_page(session: requests.Session, sosok: str, page: int) -> list[dict[str, object]]:
    response = session.get(
        f"https://finance.naver.com/sise/sise_market_sum.naver?sosok={sosok}&page={page}",
        headers={**HEADERS, "Referer": "https://finance.naver.com/"},
        timeout=30,
    )
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "lxml")
    rows: list[dict[str, object]] = []

    for tr in soup.select("table.type_2 tr"):
        link = tr.select_one("a.tltle")
        cells = tr.find_all("td")
        if link is None or len(cells) < 10:
            continue

        code_match = re.search(r"code=(\d+)", link.get("href", ""))
        if not code_match:
            continue

        code = code_match.group(1)
        name = link.get_text(strip=True)
        current_price = parse_float(cells[2].get_text(strip=True))
        change_rate_pct = parse_float(cells[4].get_text(strip=True))
        market_cap_bil_krw = parse_float(cells[6].get_text(strip=True))
        volume = parse_float(cells[9].get_text(strip=True))
        turnover_today_bil = round(current_price * volume / 100_000_000, 4)

        rows.append(
            {
                "code": code,
                "name": name,
                "market_cap_bil_krw": round(market_cap_bil_krw, 2),
                "change_rate_pct": round(change_rate_pct, 2),
                "turnover_today_bil_krw": turnover_today_bil,
            }
        )

    return rows


def fetch_naver_market_meta(session: requests.Session, market: str) -> pd.DataFrame:
    targets = ["KOSPI", "KOSDAQ"] if market == "ALL" else [market] if market in NAVER_MARKET_MAP else []
    if not targets:
        return pd.DataFrame(
            columns=["code", "market_cap_bil_krw", "change_rate_pct", "turnover_today_bil_krw"]
        )

    all_rows: list[dict[str, object]] = []
    for target in targets:
        last_page = get_naver_last_page(session, NAVER_MARKET_MAP[target])
        for page in range(1, last_page + 1):
            all_rows.extend(fetch_naver_market_page(session, NAVER_MARKET_MAP[target], page))

    return pd.DataFrame(all_rows).drop_duplicates(subset=["code"], keep="first")


def build_base_frame(session: requests.Session, start: str, end: str, market: str) -> pd.DataFrame:
    frames = [
        fetch_krx_investor_frame(session, start, end, market, code, column)
        for code, column in INVESTOR_CODES.values()
    ]

    merged = frames[0]
    for frame in frames[1:]:
        merged = merged.merge(frame, on=["code", "name"], how="outer")

    merged = merged.merge(fetch_naver_market_meta(session, market), on=["code", "name"], how="left")

    for column in [
        "private_net_buy_bil_krw",
        "trust_net_buy_bil_krw",
        "institution_net_buy_bil_krw",
        "foreigner_net_buy_bil_krw",
    ]:
        merged[column] = to_billion_krw(merged[column].fillna(0))

    merged["market_cap_bil_krw"] = merged["market_cap_bil_krw"].fillna(0).astype("float64").round(2)
    merged["change_rate_pct"] = merged["change_rate_pct"].fillna(0).astype("float64").round(2)
    merged["turnover_today_bil_krw"] = merged["turnover_today_bil_krw"].fillna(0).astype("float64").round(2)
    merged["main_net_buy_bil_krw"] = (
        merged["private_net_buy_bil_krw"] + merged["trust_net_buy_bil_krw"]
    ).round(4)

    ratio_map = {
        "private_net_buy_bil_krw": "private_ratio_pct",
        "trust_net_buy_bil_krw": "trust_ratio_pct",
        "main_net_buy_bil_krw": "main_signal_pct",
        "institution_net_buy_bil_krw": "institution_ratio_pct",
        "foreigner_net_buy_bil_krw": "foreigner_ratio_pct",
    }
    for source, target in ratio_map.items():
        merged[target] = (
            merged[source] / merged["market_cap_bil_krw"] * 100
        ).where(merged["market_cap_bil_krw"] > 0, 0).round(4)

    merged["score_pct"] = (
        merged["main_signal_pct"] * MAIN_WEIGHT
        + merged["institution_ratio_pct"] * INST_WEIGHT
        + merged["foreigner_ratio_pct"] * FOREIGN_WEIGHT
    ).round(4)

    filtered = merged[
        (merged["market_cap_bil_krw"] >= MIN_MARKET_CAP_BIL)
        & (merged["main_signal_pct"] > 0)
        & (merged["turnover_today_bil_krw"] >= MIN_TURNOVER_BIL)
        & (merged["change_rate_pct"].abs() < MAX_ABS_CHANGE_PCT)
    ].copy()

    filtered = filtered.sort_values(
        by=["score_pct", "main_signal_pct", "private_ratio_pct"],
        ascending=[False, False, False],
    ).reset_index(drop=True)
    filtered.insert(0, "rank", filtered.index + 1)
    return filtered


def add_week_over_week_deltas(
    session: requests.Session,
    frame: pd.DataFrame,
    start: str,
    end: str,
    market: str,
) -> pd.DataFrame:
    start_dt = datetime.strptime(start, "%Y%m%d")
    end_dt = datetime.strptime(end, "%Y%m%d")
    span_days = max((end_dt - start_dt).days, 0)
    prev_end = start_dt - timedelta(days=1)
    prev_start = prev_end - timedelta(days=span_days)

    previous_trust = fetch_krx_investor_frame(
        session,
        prev_start.strftime("%Y%m%d"),
        prev_end.strftime("%Y%m%d"),
        market,
        INVESTOR_CODES["trust"][0],
        "prev_trust_raw",
    )
    previous_private = fetch_krx_investor_frame(
        session,
        prev_start.strftime("%Y%m%d"),
        prev_end.strftime("%Y%m%d"),
        market,
        INVESTOR_CODES["private"][0],
        "prev_private_raw",
    )

    previous_trust["prev_trust_bil"] = to_billion_krw(previous_trust["prev_trust_raw"].fillna(0))
    previous_private["prev_private_bil"] = to_billion_krw(previous_private["prev_private_raw"].fillna(0))

    merged = frame.merge(previous_trust[["code", "prev_trust_bil"]], on="code", how="left")
    merged = merged.merge(previous_private[["code", "prev_private_bil"]], on="code", how="left")
    merged["prev_trust_bil"] = merged["prev_trust_bil"].fillna(0).astype("float64").round(4)
    merged["prev_private_bil"] = merged["prev_private_bil"].fillna(0).astype("float64").round(4)

    merged["prev_trust_ratio_pct"] = (
        merged["prev_trust_bil"] / merged["market_cap_bil_krw"] * 100
    ).where(merged["market_cap_bil_krw"] > 0, 0).round(4)
    merged["prev_private_ratio_pct"] = (
        merged["prev_private_bil"] / merged["market_cap_bil_krw"] * 100
    ).where(merged["market_cap_bil_krw"] > 0, 0).round(4)

    merged["trust_ratio_wow_change_pctp"] = (merged["trust_ratio_pct"] - merged["prev_trust_ratio_pct"]).round(4)
    merged["private_ratio_wow_change_pctp"] = (
        merged["private_ratio_pct"] - merged["prev_private_ratio_pct"]
    ).round(4)
    return merged


def compute_52week_signal_ranks(
    session: requests.Session,
    top_frame: pd.DataFrame,
    start: str,
    end: str,
    market: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    start_dt = datetime.strptime(start, "%Y%m%d")
    end_dt = datetime.strptime(end, "%Y%m%d")
    span_days = max((end_dt - start_dt).days, 0)

    top_rows = top_frame.copy()
    private_histories = {row["code"]: [] for _, row in top_rows.iterrows()}
    main_histories = {row["code"]: [] for _, row in top_rows.iterrows()}
    history_rows: list[dict[str, object]] = []

    for offset in range(52):
        window_end = end_dt - timedelta(days=7 * offset)
        window_start = window_end - timedelta(days=span_days)
        private_snapshot = fetch_krx_snapshot(
            session, window_start.strftime("%Y%m%d"), window_end.strftime("%Y%m%d"), market, INVESTOR_CODES["private"][0]
        )
        trust_snapshot = fetch_krx_snapshot(
            session, window_start.strftime("%Y%m%d"), window_end.strftime("%Y%m%d"), market, INVESTOR_CODES["trust"][0]
        )

        for _, row in top_rows.iterrows():
            code = row["code"]
            market_cap = float(row["market_cap_bil_krw"])
            private_bil = private_snapshot.get(code, 0) / 100_000_000
            trust_bil = trust_snapshot.get(code, 0) / 100_000_000

            private_ratio = round((private_bil / market_cap * 100), 4) if market_cap > 0 else 0.0
            main_ratio = round(((private_bil + trust_bil) / market_cap * 100), 4) if market_cap > 0 else 0.0

            private_histories[code].append(private_ratio)
            main_histories[code].append(main_ratio)
            history_rows.append(
                {
                    "종목코드": code,
                    "종목명": row["name"],
                    "구간시작": window_start.strftime("%Y-%m-%d"),
                    "구간종료": window_end.strftime("%Y-%m-%d"),
                    "사모 순매수(억원)": round(private_bil, 4),
                    "투신 순매수(억원)": round(trust_bil, 4),
                    "사모 비중(%)": private_ratio,
                    "메인신호(사모+투신, %)": main_ratio,
                    "주차오프셋": offset,
                }
            )

    private_rank_numbers: list[int] = []
    main_rank_numbers: list[int] = []
    private_ranks: list[str] = []
    main_ranks: list[str] = []
    for _, row in top_rows.iterrows():
        code = row["code"]
        private_values = private_histories[code]
        main_values = main_histories[code]
        private_rank = sorted(private_values, reverse=True).index(private_values[0]) + 1
        main_rank = sorted(main_values, reverse=True).index(main_values[0]) + 1
        private_rank_numbers.append(private_rank)
        main_rank_numbers.append(main_rank)
        private_ranks.append(f"{private_rank} / 52")
        main_ranks.append(f"{main_rank} / 52")

    top_rows["private_52w_rank"] = private_rank_numbers
    top_rows["main_52w_rank"] = main_rank_numbers
    top_rows["52주 내 사모신호 순위"] = private_ranks
    top_rows["52주 내 메인신호 순위"] = main_ranks
    return top_rows, pd.DataFrame(history_rows)


def finalize_display(frame: pd.DataFrame, summary: bool = False) -> pd.DataFrame:
    columns = [
        "rank",
        "code",
        "name",
        "market_cap_bil_krw",
        "private_net_buy_bil_krw",
        "trust_net_buy_bil_krw",
        "main_net_buy_bil_krw",
        "institution_net_buy_bil_krw",
        "foreigner_net_buy_bil_krw",
        "private_ratio_pct",
        "private_ratio_wow_change_pctp",
        "trust_ratio_wow_change_pctp",
        "main_signal_pct",
        "institution_ratio_pct",
        "foreigner_ratio_pct",
        "score_pct",
        "change_rate_pct",
        "turnover_today_bil_krw",
    ]
    if "52주 내 사모신호 순위" in frame.columns:
        columns += ["52주 내 사모신호 순위", "52주 내 메인신호 순위"]

    renamed = frame[columns].copy().rename(
        columns={
            "rank": "순위",
            "code": "종목코드",
            "name": "종목명",
            "market_cap_bil_krw": "시가총액(억원, 참고용)",
            "private_net_buy_bil_krw": "사모 순매수(억원, 5거래일)",
            "trust_net_buy_bil_krw": "투신 순매수(억원, 5거래일)",
            "main_net_buy_bil_krw": "사모+투신 순매수(억원, 5거래일)",
            "institution_net_buy_bil_krw": "기관 순매수(억원, 5거래일)",
            "foreigner_net_buy_bil_krw": "외국인 순매수(억원, 5거래일)",
            "private_ratio_pct": "시총 대비 사모 비중(%)",
            "private_ratio_wow_change_pctp": "사모 비중 변화(전주 대비, %p)",
            "trust_ratio_wow_change_pctp": "투신 비중 변화(전주 대비, %p)",
            "main_signal_pct": "메인신호(사모+투신, %)",
            "institution_ratio_pct": "시총 대비 기관 비중(%)",
            "foreigner_ratio_pct": "시총 대비 외국인 비중(%)",
            "score_pct": "추종 점수(%)",
            "change_rate_pct": "등락률(%)",
            "turnover_today_bil_krw": "당일 추정 거래대금(억원)",
        }
    )

    if summary:
        renamed = renamed[
            [
                "순위",
                "종목코드",
                "종목명",
                "사모+투신 순매수(억원, 5거래일)",
                "사모 비중 변화(전주 대비, %p)",
                "투신 비중 변화(전주 대비, %p)",
                "메인신호(사모+투신, %)",
                "시총 대비 기관 비중(%)",
                "시총 대비 외국인 비중(%)",
                "추종 점수(%)",
                "등락률(%)",
                "52주 내 사모신호 순위",
                "52주 내 메인신호 순위",
            ]
        ]

    return renamed


def build_readable_summary(frame: pd.DataFrame) -> pd.DataFrame:
    summary = frame[
        [
            "순위",
            "종목명",
            "사모+투신 순매수(억원, 5거래일)",
            "메인신호(사모+투신, %)",
            "사모 비중 변화(전주 대비, %p)",
            "투신 비중 변화(전주 대비, %p)",
            "시총 대비 기관 비중(%)",
            "시총 대비 외국인 비중(%)",
            "추종 점수(%)",
            "등락률(%)",
            "52주 내 사모신호 순위",
            "52주 내 메인신호 순위",
        ]
    ].copy()
    summary = summary.rename(
        columns={
            "사모+투신 순매수(억원, 5거래일)": "메인 순매수(억원)",
            "메인신호(사모+투신, %)": "메인신호(%)",
            "사모 비중 변화(전주 대비, %p)": "사모 변화(%p)",
            "투신 비중 변화(전주 대비, %p)": "투신 변화(%p)",
            "시총 대비 기관 비중(%)": "기관 비중(%)",
            "시총 대비 외국인 비중(%)": "외국인 비중(%)",
            "추종 점수(%)": "추종점수(%)",
            "52주 내 사모신호 순위": "사모 52주",
            "52주 내 메인신호 순위": "메인 52주",
        }
    )
    return summary


def build_output_path(start: str, end: str, market: str, output: str | None) -> Path:
    if output:
        return Path(output)
    return Path("output") / f"follow_rank_{market.lower()}_{start}_{end}.xlsx"


def build_json_payload(
    frame: pd.DataFrame,
    top30_with_ranks: pd.DataFrame,
    start: str,
    end: str,
    market: str,
) -> dict[str, object]:
    rows_by_code = {
        row["code"]: row
        for _, row in top30_with_ranks.iterrows()
    }
    rankings: list[dict[str, object]] = []
    for _, row in frame.iterrows():
        rank_row = rows_by_code.get(row["code"])
        rankings.append(
            {
                "rank": int(row["rank"]),
                "code": str(row["code"]),
                "name": str(row["name"]),
                "marketCapBilKrw": round(float(row["market_cap_bil_krw"]), 2),
                "privateNetBuyBilKrw": round(float(row["private_net_buy_bil_krw"]), 4),
                "trustNetBuyBilKrw": round(float(row["trust_net_buy_bil_krw"]), 4),
                "mainNetBuyBilKrw": round(float(row["main_net_buy_bil_krw"]), 4),
                "institutionNetBuyBilKrw": round(float(row["institution_net_buy_bil_krw"]), 4),
                "foreignerNetBuyBilKrw": round(float(row["foreigner_net_buy_bil_krw"]), 4),
                "privateRatioPct": round(float(row["private_ratio_pct"]), 4),
                "privateRatioWowChangePctp": round(float(row["private_ratio_wow_change_pctp"]), 4),
                "trustRatioWowChangePctp": round(float(row["trust_ratio_wow_change_pctp"]), 4),
                "mainSignalPct": round(float(row["main_signal_pct"]), 4),
                "institutionRatioPct": round(float(row["institution_ratio_pct"]), 4),
                "foreignerRatioPct": round(float(row["foreigner_ratio_pct"]), 4),
                "scorePct": round(float(row["score_pct"]), 4),
                "changeRatePct": round(float(row["change_rate_pct"]), 2),
                "turnoverTodayBilKrw": round(float(row["turnover_today_bil_krw"]), 2),
                "private52wRank": int(rank_row["private_52w_rank"]) if rank_row is not None else None,
                "main52wRank": int(rank_row["main_52w_rank"]) if rank_row is not None else None,
                "naverFinanceUrl": f"https://finance.naver.com/item/main.naver?code={str(row['code'])}",
            }
        )

    top_score_row = max(rankings, key=lambda item: float(item["scorePct"]))
    top_signal_row = max(rankings, key=lambda item: float(item["mainSignalPct"]))
    top_trust_delta_row = max(rankings, key=lambda item: float(item["trustRatioWowChangePctp"]))
    top_52w_count = sum(1 for item in rankings[:30] if item["main52wRank"] == 1)
    average_signal = round(sum(float(item["mainSignalPct"]) for item in rankings) / max(len(rankings), 1), 4)

    return {
        "overview": {
            "title": "사모·투신 추종 랭킹",
            "subtitle": "최근 5거래일 기준 사모+투신 유입을 시가총액 대비 비율로 비교한 화면입니다.",
            "asOfDate": f"{end[:4]}-{end[4:6]}-{end[6:8]}",
            "dateRange": {
                "start": f"{start[:4]}-{start[4:6]}-{start[6:8]}",
                "end": f"{end[:4]}-{end[4:6]}-{end[6:8]}",
            },
            "marketLabel": f"{market} / KRX 기준",
            "unitLabel": "단위: 억원 / %",
        },
        "summary": {
            "totalTracked": len(rankings),
            "topScoreName": top_score_row["name"],
            "topScorePct": top_score_row["scorePct"],
            "topMainSignalName": top_signal_row["name"],
            "topMainSignalPct": top_signal_row["mainSignalPct"],
            "topTrustAccelerationName": top_trust_delta_row["name"],
            "topTrustAccelerationPctp": top_trust_delta_row["trustRatioWowChangePctp"],
            "averageMainSignalPct": average_signal,
            "main52wTopCount": top_52w_count,
        },
        "rankings": rankings,
        "generatedAt": datetime.now().isoformat(),
    }


def format_sheet(worksheet, summary: bool = False) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 8,
        "B": 12,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 20,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 16,
        "L": 16,
        "M": 14,
        "N": 12,
        "O": 16,
        "P": 16,
        "Q": 18,
        "R": 18,
        "S": 18,
        "T": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, worksheet.max_row + 1):
        for cell in worksheet[row]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    if "종목명" in [cell.value for cell in worksheet[1]]:
        name_col = {cell.value: cell.column_letter for cell in worksheet[1]}["종목명"]
        for row in range(2, worksheet.max_row + 1):
            worksheet[f"{name_col}{row}"].alignment = Alignment(horizontal="left", vertical="center")

    if summary:
        score_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
        signal_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
        rank_fill = PatternFill(fill_type="solid", fgColor="DCFCE7")
        header_map = {cell.value: cell.column_letter for cell in worksheet[1]}
        for header in ["메인신호(사모+투신, %)", "추종 점수(%)"]:
            if header in header_map:
                worksheet[f"{header_map[header]}1"].fill = score_fill if "점수" in header else signal_fill
                worksheet[f"{header_map[header]}1"].font = Font(bold=True)
        for header in ["52주 내 사모신호 순위", "52주 내 메인신호 순위"]:
            if header in header_map:
                worksheet[f"{header_map[header]}1"].fill = rank_fill
                worksheet[f"{header_map[header]}1"].font = Font(bold=True)

        for row_idx in range(2, worksheet.max_row + 1):
            if row_idx <= 4:
                fill = PatternFill(fill_type="solid", fgColor="FFF7ED")
                for cell in worksheet[row_idx]:
                    cell.fill = fill

        for target_header in ["메인신호(%)", "추종점수(%)", "메인신호(사모+투신, %)", "추종 점수(%)"]:
            if target_header in header_map:
                col = header_map[target_header]
                for row_idx in range(2, worksheet.max_row + 1):
                    worksheet[f"{col}{row_idx}"].font = Font(bold=True)

    percent_headers = {
        "시총 대비 사모 비중(%)",
        "사모 비중 변화(전주 대비, %p)",
        "투신 비중 변화(전주 대비, %p)",
        "메인신호(사모+투신, %)",
        "메인신호(%)",
        "시총 대비 기관 비중(%)",
        "기관 비중(%)",
        "시총 대비 외국인 비중(%)",
        "외국인 비중(%)",
        "추종 점수(%)",
        "추종점수(%)",
        "등락률(%)",
        "사모 변화(%p)",
        "투신 변화(%p)",
    }
    header_map = {cell.value: cell.column_letter for cell in worksheet[1]}
    for header in percent_headers:
        if header in header_map:
            col = header_map[header]
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet[f"{col}{row_idx}"].number_format = '0.0000"%"'

    amount_headers = {
        "시가총액(억원, 참고용)",
        "사모 순매수(억원, 5거래일)",
        "투신 순매수(억원, 5거래일)",
        "사모+투신 순매수(억원, 5거래일)",
        "기관 순매수(억원, 5거래일)",
        "외국인 순매수(억원, 5거래일)",
        "당일 추정 거래대금(억원)",
        "메인 순매수(억원)",
        "사모 순매수(억원)",
        "투신 순매수(억원)",
    }
    for header in amount_headers:
        if header in header_map:
            col = header_map[header]
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet[f"{col}{row_idx}"].number_format = '#,##0.0000'


def create_dashboard_sheet(writer, readable_frame: pd.DataFrame, full_frame: pd.DataFrame, start: str, end: str) -> None:
    book = writer.book
    ws = book.create_sheet("대시보드", 0)
    dark_fill = PatternFill(fill_type="solid", fgColor="111827")
    panel_fill = PatternFill(fill_type="solid", fgColor="172033")
    header_fill = PatternFill(fill_type="solid", fgColor="2B354A")
    accent_fill = PatternFill(fill_type="solid", fgColor="D8A64F")
    cyan_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    green_fill = PatternFill(fill_type="solid", fgColor="14532D")
    border = Border(
        left=Side(style="thin", color="314158"),
        right=Side(style="thin", color="314158"),
        top=Side(style="thin", color="314158"),
        bottom=Side(style="thin", color="314158"),
    )

    for row in range(1, 45):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.fill = dark_fill

    widths = {
        "A": 10, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16,
        "H": 16, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16, "N": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[14].height = 24

    ws.merge_cells("A2:J6")
    hero = ws["A2"]
    hero.value = "PRIVATE EQUITY FLOW MONITOR\n52주 순매수 랭킹"
    hero.fill = panel_fill
    hero.font = Font(color="FFFFFF", bold=True, size=24)
    hero.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row in ws["A2:J6"]:
        for cell in row:
            cell.fill = panel_fill
            cell.border = border

    ws["A7"] = "사모+투신 유입을 메인으로, 기관/외인을 보조 지표로 보는 추종 랭킹입니다."
    ws["A7"].font = Font(color="B8C4D9", size=11)

    ws.merge_cells("K2:N6")
    stamp = ws["K2"]
    stamp.value = f"KRX 기준\n{end[:4]}-{end[4:6]}-{end[6:]}\n조회 구간 {start[:4]}-{start[4:6]}-{start[6:]} ~ {end[:4]}-{end[4:6]}-{end[6:]}"
    stamp.fill = panel_fill
    stamp.font = Font(color="FFFFFF", bold=True, size=16)
    stamp.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws["K2:N6"]:
        for cell in row:
            cell.fill = panel_fill
            cell.border = border

    top_row = full_frame.iloc[0]
    best_signal_row = full_frame.sort_values("메인신호(사모+투신, %)", ascending=False).iloc[0]
    avg_signal = round(float(full_frame["메인신호(사모+투신, %)"].mean()), 4)
    rank1_count = int((readable_frame["메인 52주"] == "1 / 52").sum())

    cards = [
        ("A9:C12", "표본 종목 수", str(len(full_frame)), "조건 필터 통과 기준"),
        ("D9:F12", "추종 점수 1위", str(top_row["종목명"]), f'{top_row["추종 점수(%)"]:.4f}%'),
        ("G9:I12", "메인신호 최고", str(best_signal_row["종목명"]), f'{best_signal_row["메인신호(사모+투신, %)"]:.4f}%'),
        ("J9:L12", "평균 메인신호", f"{avg_signal:.4f}%", "현재 표본 평균"),
        ("M9:N12", "52주 메인 1위", f"{rank1_count}개", "상위30 기준"),
    ]

    for ref, label, value, sub in cards:
        ws.merge_cells(ref)
        start_cell = ref.split(":")[0]
        cell = ws[start_cell]
        cell.value = f"{label}\n{value}\n{sub}"
        cell.fill = panel_fill
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row in ws[ref]:
            for item in row:
                item.fill = panel_fill
                item.border = border

    ws["A14"] = "핵심 추종 테이블"
    ws["A14"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A15"] = "정렬 기준: 추종점수 내림차순 / 메인신호 양수 / 시총 3000억 이상"
    ws["A15"].font = Font(color="B8C4D9", size=10)

    dashboard_headers = [
        "순위",
        "종목명",
        "메인 순매수(억원)",
        "메인신호(%)",
        "사모 변화(%p)",
        "투신 변화(%p)",
        "기관 비중(%)",
        "외국인 비중(%)",
        "추종점수(%)",
        "등락률(%)",
        "사모 52주",
        "메인 52주",
    ]
    start_row = 17
    for col_idx, header in enumerate(dashboard_headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_offset, row in enumerate(readable_frame.head(12).itertuples(index=False), start=1):
        excel_row = start_row + row_offset
        values = list(row)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = panel_fill
            cell.font = Font(color="FFFFFF", bold=(col_idx in {2, 4, 9}))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_idx in {4, 5, 6, 7, 8, 9, 10}:
                cell.number_format = '0.0000"%"'
            if col_idx == 3:
                cell.number_format = '#,##0.0000'

        if excel_row <= start_row + 3:
            for col_idx in range(1, len(dashboard_headers) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = accent_fill
                ws.cell(row=excel_row, column=col_idx).font = Font(color="111827", bold=True)

        if values[9] is not None:
            change_cell = ws.cell(row=excel_row, column=10)
            change_cell.fill = green_fill if float(values[9]) < 0 else cyan_fill

    ws.freeze_panes = "A17"


def save_excel(
    output_path: Path,
    readable_frame: pd.DataFrame,
    top30_frame: pd.DataFrame,
    full_frame: pd.DataFrame,
    history_frame: pd.DataFrame,
    start: str,
    end: str,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        create_dashboard_sheet(writer, readable_frame, full_frame, start, end)
        readable_frame.to_excel(writer, index=False, sheet_name="핵심30")
        top30_frame.to_excel(writer, index=False, sheet_name="상위30")
        full_frame.to_excel(writer, index=False, sheet_name="전체랭킹")
        history_frame.to_excel(writer, index=False, sheet_name="52주이력")

        format_sheet(writer.sheets["핵심30"], summary=True)
        format_sheet(writer.sheets["상위30"], summary=True)
        format_sheet(writer.sheets["전체랭킹"], summary=False)
        format_sheet(writer.sheets["52주이력"], summary=False)

        info = writer.book.create_sheet("설명")
        info["A1"] = "기준"
        info["A2"] = f"조회 구간: {start} ~ {end}"
        info["A3"] = "메인신호 = (사모 5거래일 순매수 + 투신 5거래일 순매수) / 시가총액 * 100"
        info["A4"] = "추종 점수 = 메인신호*0.85 + 기관비중*0.10 + 외국인비중*0.05"
        info["A5"] = "필터 = 시총 3000억 이상, 메인신호 양수, 당일 추정 거래대금 30억 이상, 등락률 절대값 25% 미만"
        info["A6"] = "52주 순위는 같은 길이 구간을 7일씩 뒤로 밀어 최근 52개 구간과 비교한 순위"
        info["A7"] = "시가총액은 참고용이며 최근 공개 기준 값이 들어갈 수 있습니다."
        info["A8"] = "추천 사용법: 핵심30 시트만 먼저 보고, 필요할 때 상위30/전체랭킹에서 세부 숫자를 확인하세요."
        info.column_dimensions["A"].width = 120


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a follow-ranking Excel report.")
    parser.add_argument("--start", help="Start date in YYYYMMDD format")
    parser.add_argument("--end", help="End date in YYYYMMDD format")
    parser.add_argument("--date", help="Anchor date in YYYYMMDD format")
    parser.add_argument("--market", default="ALL", choices=["KOSPI", "KOSDAQ", "KONEX", "ALL"])
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--output")
    parser.add_argument("--open", action="store_true")
    parser.add_argument("--json", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    session = requests.Session()

    if args.start and args.end:
        start, end = args.start, args.end
    elif args.date:
        anchor = args.date.replace("-", "")
        start, end = get_recent_trading_window(session, anchor)
    elif args.end:
        anchor = args.end.replace("-", "")
        start, end = get_recent_trading_window(session, anchor)
    else:
        start, end = get_recent_trading_window(session)

    base_frame = build_base_frame(session, start, end, args.market)
    base_frame = add_week_over_week_deltas(session, base_frame, start, end, args.market)
    top30_raw = base_frame.head(30).copy()
    top30_with_ranks, history_frame = compute_52week_signal_ranks(session, top30_raw, start, end, args.market)

    if args.json:
        payload = build_json_payload(base_frame.head(args.limit).copy(), top30_with_ranks, start, end, args.market)
        print(json.dumps(payload, ensure_ascii=False))
        return

    top30_display = finalize_display(top30_with_ranks, summary=True)
    readable_display = build_readable_summary(top30_display)
    full_display = finalize_display(base_frame.head(args.limit).copy(), summary=False)

    output_path = build_output_path(start, end, args.market, args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.suffix.lower() == ".csv":
        full_display.to_csv(output_path, index=False, encoding="utf-8-sig")
    else:
        save_excel(output_path, readable_display, top30_display, full_display, history_frame, start, end)

    print(f"saved: {output_path.resolve()}")
    print(top30_display.head(20).to_string(index=False))

    if args.open:
        os.startfile(output_path.resolve())


if __name__ == "__main__":
    main()
